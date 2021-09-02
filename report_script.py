#!/usr/bin/env python3

import sys
import os
import io
import urllib3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.drawing.image import Image
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint


r = 1
http = urllib3.PoolManager()
url = 'https://www.silicon-mobility.com/silicon-mobility/wp-content/'\
      'themes/silicon-mobility/dist/images/SM_logo_color.png'

r = http.request('GET', url)
image_file = io.BytesIO(r.data)
img = Image(image_file)

myset = set()

# *************defining borders configuration*****************
thin_border = Border(
    left=Side(border_style=BORDER_THIN, color='00339966'),
    right=Side(border_style=BORDER_THIN, color='00339966'),
    top=Side(border_style=BORDER_THIN, color='00339966'),
    bottom=Side(border_style=BORDER_THIN, color='00339966')
    )

thin_border_ws1 = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000')
    )

double_border = Border(
    left=Side(border_style="double", color='00339966'),
    right=Side(border_style="double", color='00339966'),
    top=Side(border_style="double", color='00339966'),
    bottom=Side(border_style="double", color='00339966')
    )
# ********************************************************


def fn1(fileName, workbook, ws2):
    # *********************styling part**********************
    # cells colors to be filled with for diagnostic type cell field
    redFill = PatternFill(fgColor='ee2424', fill_type='solid')
    yellowFill = PatternFill(fgColor="ffff9e", fill_type='solid')

    # configuring gridlines
    ws2.sheet_view.showGridLines = False
    # adding the entreprise logo
    ws2.add_image(img, 'H1')
    # Sheet Title
    ws2.merge_cells('B2:G8')
    ws2["B2"] = "AUTOSAR C++14 rules compliance results"
    ws2["B2"].font = Font(size=35, bold=True, color='00008000')
    ws2["B2"].alignment = Alignment(
            horizontal='center', vertical='center', wrap_text=False
            )

    # Document description part
    ws2['B10'] = "Document description  : "
    ws2["B10"].alignment = Alignment(vertical='top', wrap_text=True)
    ws2.merge_cells('B10:C13')
    ws2["B10"].border = double_border
    ws2["B11"].border = double_border
    ws2["B12"].border = double_border
    ws2["B13"].border = double_border
    ws2["C10"].border = double_border
    ws2["C11"].border = double_border
    ws2["C12"].border = double_border
    ws2["C13"].border = double_border

    # Header part
    ws2.merge_cells('B15:C15')
    ws2['B15'] = "Header sheet"
    ws2["B15"].fill = PatternFill(fgColor='138457', fill_type='solid')
    ws2['B15'].border = thin_border
    ws2["B15"].alignment = Alignment(horizontal='center')
    ws2["B15"].font = Font(bold=True, color='FFFFFFFF', underline="single")
    ws2['B16'] = "Project Path"
    ws2["B16"].fill = PatternFill(fgColor='70ad47', fill_type='solid')
    ws2["B16"].border = thin_border
    ws2["B16"].alignment = Alignment(horizontal='center')
    ws2["B16"].font = Font(bold=True, color='FFFFFFFF')
    ws2['B17'] = "Project Name"
    ws2["B17"].fill = PatternFill(fgColor='70ad47', fill_type='solid')
    ws2["B17"].border = thin_border
    ws2["B17"].alignment = Alignment(horizontal='center')
    ws2["B17"].font = Font(bold=True, color='FFFFFFFF')
    ws2['B18'] = "Release"
    ws2["B18"].fill = PatternFill(fgColor='70ad47', fill_type='solid')
    ws2["B18"].border = thin_border
    ws2["B18"].alignment = Alignment(horizontal='center')
    ws2["B18"].font = Font(bold=True, color='FFFFFFFF')
    ws2["C15"].border = thin_border
    ws2["C16"].border = thin_border
    ws2["C17"].border = thin_border
    ws2['C18'] = "R2020R2"
    ws2["C18"].border = thin_border

    # Main table part
    ws2.merge_cells('B20:G20')
    ws2["B20"] = "Device Utilization Sheet"
    ws2["B20"].fill = PatternFill(fgColor='138457', fill_type='solid')
    ws2["B20"].alignment = Alignment(horizontal='center')
    ws2["B20"].font = Font(bold=True, color='FFFFFFFF', underline="single")
    ws2['B20'].border = thin_border
    ws2['C20'].border = thin_border
    ws2['D20'].border = thin_border
    ws2['E20'].border = thin_border
    ws2['F20'].border = thin_border
    ws2['G20'].border = thin_border
    ws2["B21"] = "file path"
    ws2['B21'].border = thin_border
    ws2["B21"].fill = PatternFill(fgColor='70ad47', fill_type='solid')
    ws2["C21"] = "Line"
    ws2['C21'].border = thin_border
    ws2["C21"].fill = PatternFill(fgColor='70ad47', fill_type='solid')
    ws2["D21"] = "Column"
    ws2['D21'].border = thin_border
    ws2["D21"].fill = PatternFill(fgColor='70ad47', fill_type='solid')
    ws2["E21"] = "Severity"
    ws2['E21'].border = thin_border
    ws2["E21"].fill = PatternFill(fgColor='70ad47', fill_type='solid')
    ws2["F21"] = "diagnostic description"
    ws2['F21'].border = thin_border
    ws2["F21"].fill = PatternFill(fgColor='70ad47', fill_type='solid')
    ws2["G21"] = "diagnostic rule"
    ws2['G21'].border = thin_border
    ws2["G21"].fill = PatternFill(fgColor='70ad47', fill_type='solid')
    ws2["B21"].font = Font(bold=True, color='FFFFFFFF')
    ws2["C21"].font = Font(bold=True, color='FFFFFFFF')
    ws2["D21"].font = Font(bold=True, color='FFFFFFFF')
    ws2["E21"].font = Font(bold=True, color='FFFFFFFF')
    ws2["F21"].font = Font(bold=True, color='FFFFFFFF')
    ws2["G21"].font = Font(bold=True, color='FFFFFFFF')

    # defining width for columns
    ws2.column_dimensions['B'].width = 55
    ws2.column_dimensions['F'].width = 55
    ws2.column_dimensions['G'].width = 25
    count = 20
    f = open(fileName, 'r')
    while (True):
        line = f.readline()
        if not line:
            break
        else:
            # pick the lines where there is a warning or an error
            if line.find("warning:") >= 0:
                fn2(line, "warning:", count, ws2, yellowFill)
                count += 1
            elif line.find("error:") >= 0:
                fn2(line, "error:", count, ws2, redFill)
                count += 1
    f.close()
    # The excel file will be created in the same directory of
    # the text file and with the same name
    workbook.save(fileName.replace(".txt", ".xlsx"))


def fn2(lineOfTreatement, DiagnosticType, count, ws2, diagnosticTypeCellcolor):
    """This function will treat each line containig an error/warning
    """
    listOfElements = lineOfTreatement.split(DiagnosticType)
    place_line_column = listOfElements[0].split(':')
    place = place_line_column[0]
    line = place_line_column[-3]
    column = place_line_column[-2]
    diagnosticType = DiagnosticType.replace(':', '')
    diagnosticTextInList = listOfElements[1].split()
    diagnosticText = ""
    for element in diagnosticTextInList[:-1]:
        diagnosticText += element + ' '
    rule = diagnosticTextInList[-1].replace(']', '').replace('[', '')

    # Add distinct references
    myset.add(rule)

    ws2["B" + str(count+2)] = place
    ws2["C" + str(count+2)] = int(line)
    ws2["D" + str(count+2)] = int(column)
    ws2["E" + str(count+2)] = diagnosticType
    ws2["F" + str(count+2)] = diagnosticText
    ws2["G" + str(count+2)] = rule

    # *********************styling part**********************
    # defining cells borders
    ws2["B" + str(count+2)].border = thin_border
    ws2["C" + str(count+2)].border = thin_border
    ws2["D" + str(count+2)].border = thin_border
    ws2["E" + str(count+2)].border = thin_border
    ws2["F" + str(count+2)].border = thin_border
    ws2["G" + str(count+2)].border = thin_border
    # make the text wrap automatically
    ws2["B" + str(count+2)].alignment = Alignment(wrap_text=True)
    ws2["F" + str(count+2)].alignment = Alignment(wrap_text=True)
    ws2["G" + str(count+2)].alignment = Alignment(wrap_text=True)
    # fill the cell background with the appropriate color to describe
    # if it is a warning(yellow) or an error(red)
    ws2["E" + str(count+2)].fill = diagnosticTypeCellcolor
    # ********************************************************


def draw_chart(filename, workbook, sheet, tmp_sheet):
    total_rules = 342
    nbr_failed_tests = len(myset)
    nbr_passed_tests = total_rules - nbr_failed_tests
    data = [
        ['Labels', 'Values'],
        ['Passed Tests', nbr_passed_tests / 100],
        ['Failed Tests', nbr_failed_tests / 100],
        ['N/A', 0.00001],
    ]

    for row in data:
        tmp_sheet.append(row)

    pie = PieChart()
    labels = Reference(tmp_sheet, min_col=1, min_row=2, max_row=4)
    data = Reference(tmp_sheet, min_col=2, min_row=1, max_row=4)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Tests Result"

    # show data labels as percentage
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    pie.dataLabels.showVal = False
    pie.dataLabels.showCatName = False

    # set colors of the chart
    series = pie.series[0]
    pt0 = DataPoint(idx=0)
    pt0.graphicalProperties.solidFill = "70ad47"
    pt1 = DataPoint(idx=1)
    pt1.graphicalProperties.solidFill = "ee2424"
    pt2 = DataPoint(idx=2)
    pt2.graphicalProperties.solidFill = "ffff9e"
    series.dPt.append(pt0)
    series.dPt.append(pt1)
    series.dPt.append(pt2)

    sheet.add_chart(pie, "I12")
    workbook.save(filename.replace(".txt", ".xlsx"))


def historySheetStyle(ws1):
    """First Sheet
    """
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions['A'].width = 55
    ws1.column_dimensions['B'].width = 70
    ws1["A1"] = "Document Title"
    ws1['A1'].border = thin_border_ws1
    ws1["A1"].fill = PatternFill(fgColor='d0cece', fill_type='solid')
    ws1["A1"].font = Font(bold=True)
    ws1["B1"] = "AUTOSAR C++14 rules compliance report"
    ws1['B1'].border = thin_border_ws1
    ws1["A2"] = "Document Responsibility"
    ws1['A2'].border = thin_border_ws1
    ws1["A2"].font = Font(bold=True)
    ws1["A2"].fill = PatternFill(fgColor='d0cece', fill_type='solid')
    ws1['B2'].border = thin_border_ws1
    ws1["A3"] = "Document Status"
    ws1['A3'].border = thin_border_ws1
    ws1["A3"].font = Font(bold=True)
    ws1["A3"].fill = PatternFill(fgColor='d0cece', fill_type='solid')
    ws1["B3"] = "<Draft | Ready for Approval | Final>"
    ws1['B3'].border = thin_border_ws1
    ws1["A4"] = "Author"
    ws1['A4'].border = thin_border_ws1
    ws1["A4"].fill = PatternFill(fgColor='d0cece', fill_type='solid')
    ws1["A4"].font = Font(bold=True)
    ws1['B4'].border = thin_border_ws1
    ws1["A5"] = "Reviewer"
    ws1['A5'].border = thin_border_ws1
    ws1["A5"].fill = PatternFill(fgColor='d0cece', fill_type='solid')
    ws1["A5"].font = Font(bold=True)
    ws1['B5'].border = thin_border_ws1
    ws1["A8"] = "Projet Name"
    ws1['A8'].border = thin_border_ws1
    ws1["A8"].fill = PatternFill(fgColor='d0cece', fill_type='solid')
    ws1["A8"].font = Font(bold=True)
    ws1["B8"] = "Autosar Project"
    ws1['B8'].border = thin_border_ws1
    ws1["A9"] = "Release"
    ws1['A9'].border = thin_border
    ws1["A9"].fill = PatternFill(fgColor='d0cece', fill_type='solid')
    ws1["A9"].font = Font(bold=True)
    ws1["B9"] = "R2020R2"
    ws1['B9'].border = thin_border_ws1
    ws1["A10"] = "Path_To_Add"
    ws1['A10'].border = thin_border_ws1
    ws1["A10"].font = Font(bold=True)
    ws1["A10"].fill = PatternFill(fgColor='d0cece', fill_type='solid')
    ws1["B10"] = "C:\\Users\\rjeridi\\Desktop\\Scripts Archtiecture"
    ws1['B10'].border = thin_border_ws1
    ws1["A11"] = "Board"
    ws1['A11'].border = thin_border_ws1
    ws1["A11"].fill = PatternFill(fgColor='d0cece', fill_type='solid')
    ws1["A11"].font = Font(bold=True)
    ws1["B11"] = "Not specified"
    ws1['B11'].border = thin_border_ws1


def pathOfFile(path):
    if path.find('/') >= 0:
        path = path.replace("/", "//")
    return path


def main():
    filename = pathOfFile(sys.argv[1])
    workbook = Workbook()
    ws1 = workbook.active
    ws2 = workbook.create_sheet()
    ws1 = workbook.worksheets[0]
    ws1.title = 'Sheet_History'
    ws2 = workbook.worksheets[1]
    ws2.title = 'Main_sheet'
    tmp_sheet = workbook.create_sheet(title="temporary file")
    historySheetStyle(ws1)
    fn1(filename, workbook, ws2)
    draw_chart(filename, workbook, ws2, tmp_sheet)


if __name__ == "__main__":
    main()
